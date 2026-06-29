from __future__ import annotations

from io import BytesIO
from typing import Any
import math

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.graphics.shapes import Drawing, Rect, String, Line, Circle, Ellipse, Path


def build_excel(results: dict[str, Any]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    ws.append(["Output", "Value", "Unit"])
    for row in results.get("summary_rows", []):
        ws.append([row.get("output"), row.get("value"), row.get("unit")])

    ws2 = wb.create_sheet("Data Waktu")
    ts = results.get("time_series", [])
    if ts:
        headers = list(ts[0].keys())
        ws2.append(headers)
        for item in ts:
            ws2.append([item.get(h) for h in headers])

    ws3 = wb.create_sheet("Visualisasi Series")
    vs = results.get("visualization_series", [])
    if vs:
        headers = list(vs[0].keys())
        ws3.append(headers)
        for item in vs:
            ws3.append([item.get(h) for h in headers])

    conv_ts = results.get("conventional_time_series", [])
    if conv_ts:
        wsconv = wb.create_sheet("Data Waktu Konvensional")
        headers = list(conv_ts[0].keys())
        wsconv.append(headers)
        for item in conv_ts:
            wsconv.append([item.get(h) for h in headers])

    cond_rows = results.get("condition_rows", [])
    if cond_rows:
        ws4 = wb.create_sheet("Kondisi Refrigeran")
        headers = list(cond_rows[0].keys())
        ws4.append(headers)
        for item in cond_rows:
            ws4.append([item.get(h) for h in headers])

    pd_rows = results.get("pressure_drop_rows", [])
    if pd_rows:
        wspd = wb.create_sheet("Pressure Drop")
        headers = list(pd_rows[0].keys())
        wspd.append(headers)
        for item in pd_rows:
            wspd.append([item.get(h) for h in headers])

    ht_rows = results.get("heat_transfer_rows", [])
    if ht_rows:
        wsht = wb.create_sheet("Perpindahan Panas")
        headers = list(ht_rows[0].keys())
        wsht.append(headers)
        for item in ht_rows:
            wsht.append([item.get(h) for h in headers])

    ass_rows = results.get("assumption_rows", [])
    if ass_rows:
        wsas = wb.create_sheet("Asumsi Model")
        headers = list(ass_rows[0].keys())
        wsas.append(headers)
        for item in ass_rows:
            wsas.append([item.get(h) for h in headers])

    ws5 = wb.create_sheet("Geometri")
    ws5.append(["Komponen", "Parameter", "Value", "Unit"])
    for row in results.get("geometry_rows", []):
        ws5.append([row.get("komponen"), row.get("parameter"), row.get("value"), row.get("unit")])

    ws6 = wb.create_sheet("Persentase Coil HWST")
    coil = results.get("coil_usage", [])
    if coil:
        headers = ["sistem", "komponen", "zona", "persentase_COP_best_DSH", "persentase_akhir", "t_COP_best_DSH_min", "T_tank_COP_best_DSH_C", "catatan"]
        ws6.append(headers)
        for item in coil:
            ws6.append([item.get(h) for h in headers])

    ws6b = wb.create_sheet("Persentase Coil Konv")
    coil_base = results.get("coil_usage_conventional", [])
    if coil_base:
        headers = ["sistem", "komponen", "zona", "persentase_steady", "catatan"]
        ws6b.append(headers)
        for item in coil_base:
            ws6b.append([item.get(h) for h in headers])

    valid_rows = results.get("validation_rows", [])
    if valid_rows:
        wsval = wb.create_sheet("Validasi Siklus")
        headers = list(valid_rows[0].keys())
        wsval.append(headers)
        for item in valid_rows:
            wsval.append([item.get(h) for h in headers])

    ph = results.get("ph_series", [])
    if ph:
        ws7 = wb.create_sheet("PH_State")
        ws7.append(["time_min", "label", "name", "h_kJ_kg", "P_kPa", "T_C"])
        for frame in ph:
            for pt in frame.get("points", []):
                ws7.append([frame.get("time_min"), pt.get("label"), pt.get("name"), pt.get("h_kJ_kg"), pt.get("P_kPa"), pt.get("T_C")])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 12
            letter = col[0].column_letter
            for cell in col:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            sheet.column_dimensions[letter].width = min(max_len + 2, 42)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _f(value: Any, nd: int = 2) -> str:
    try:
        return f"{float(value):.{nd}f}"
    except Exception:
        return "-"


def _basic_table_style(font_size: int = 8) -> TableStyle:
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#94a3b8")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
    ])


def _latest_row(results: dict[str, Any]) -> dict[str, Any]:
    ts = results.get("visualization_series") or results.get("time_series") or []
    return ts[min(len(ts)-1, max(0, len(ts)//2))] if ts else {}


def _temp_color(t: Any):
    try:
        v = max(0.0, min(1.0, float(t) / 90.0))
    except Exception:
        v = 0.5
    if v < 0.25:
        return colors.HexColor("#0891b2")
    if v < 0.5:
        return colors.HexColor("#22c55e")
    if v < 0.75:
        return colors.HexColor("#f59e0b")
    return colors.HexColor("#dc2626")


def _build_system_drawing(results: dict[str, Any]) -> Drawing:
    row = _latest_row(results)
    d = Drawing(520, 250)
    d.add(Rect(0, 0, 520, 250, rx=14, ry=14, fillColor=colors.HexColor("#f8fafc"), strokeColor=colors.HexColor("#cbd5e1")))
    comp=(55,155); tank=(210,164); cond=(420,165); cap=(340,75); evap=(145,75)
    def pipe(points, temp_key):
        p = Path(); p.moveTo(points[0][0], points[0][1])
        for x,y in points[1:]: p.lineTo(x,y)
        p.strokeColor = _temp_color(row.get(temp_key)); p.strokeWidth = 8; p.fillColor = None
        d.add(p)
    pipe([(comp[0]+42, comp[1]), (tank[0]-92, tank[1])], "T_comp_out_C")
    pipe([(tank[0]+105, tank[1]), (cond[0]-55, cond[1])], "T_ref_out_HWST_C")
    pipe([(cond[0], cond[1]-45), (cond[0], cap[1]+10), (cap[0]+40, cap[1])], "T_cond_out_C")
    pipe([(cap[0]-40, cap[1]), (evap[0]+70, evap[1])], "T_evap_in_C")
    pipe([(evap[0]-70, evap[1]), (comp[0]-20, evap[1]), (comp[0]-20, comp[1]-35)], "T_evap_out_C")
    d.add(Ellipse(comp[0], comp[1], 35, 50, fillColor=colors.white, strokeColor=colors.HexColor("#1e293b"), strokeWidth=1.5))
    d.add(String(comp[0]-28, comp[1]-72, "KOMPRESOR", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(Rect(tank[0]-90, tank[1]-30, 180, 60, rx=28, ry=28, fillColor=colors.white, strokeColor=colors.HexColor("#1e293b"), strokeWidth=1.5))
    for i in range(8):
        d.add(Line(tank[0]-65+i*17, tank[1]-18, tank[0]-65+i*17, tank[1]+18, strokeColor=colors.HexColor("#f59e0b"), strokeWidth=2.2))
    d.add(String(tank[0]-20, tank[1]-52, "HWST horizontal", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(Rect(cond[0]-45, cond[1]-50, 90, 100, rx=10, ry=10, fillColor=colors.white, strokeColor=colors.HexColor("#1e293b"), strokeWidth=1.5))
    for i in range(7):
        d.add(Line(cond[0]-30+i*10, cond[1]-35, cond[0]-30+i*10, cond[1]+35, strokeColor=colors.HexColor("#2563eb"), strokeWidth=2))
    d.add(String(cond[0]-28, cond[1]-70, "KONDENSOR", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(Rect(evap[0]-65, evap[1]-30, 130, 60, rx=10, ry=10, fillColor=colors.white, strokeColor=colors.HexColor("#1e293b"), strokeWidth=1.5))
    for i in range(5):
        d.add(Line(evap[0]-48, evap[1]-18+i*8, evap[0]+48, evap[1]-18+i*8, strokeColor=colors.HexColor("#2563eb"), strokeWidth=2.2))
    d.add(String(evap[0]-36, evap[1]-48, "EVAPORATOR", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(Circle(cap[0], cap[1], 13, fillColor=colors.HexColor("#f59e0b"), strokeColor=colors.HexColor("#92400e"), strokeWidth=1.5))
    d.add(String(cap[0]-20, cap[1]-30, "KAPILER", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(String(25, 225, f"T kompresor out: {_f(row.get('T_comp_out_C'),1)} °C", fontSize=8, fillColor=colors.HexColor("#dc2626")))
    d.add(String(25, 211, f"T keluar HWST: {_f(row.get('T_ref_out_HWST_C'),1)} °C", fontSize=8, fillColor=colors.HexColor("#ea580c")))
    d.add(String(25, 197, f"T air tangki: {_f(row.get('T_tank_mean_C'),1)} °C | Q HWST: {_f(row.get('Q_HWST_kW'),2)} kW", fontSize=8, fillColor=colors.HexColor("#334155")))
    return d


def _build_ph_drawing(results: dict[str, Any]) -> Drawing:
    frames = results.get("ph_series", [])
    frame = frames[min(len(frames)-1, max(0, len(frames)//2))] if frames else {"points": []}
    dome = results.get("ph_dome", {}) or {}
    pts = []
    pts += dome.get("liquid", [])[:]
    pts += dome.get("vapor", [])[:]
    pts += frame.get("points", [])
    vals = [(float(p.get("h_kJ_kg")), float(p.get("P_kPa"))) for p in pts if p.get("h_kJ_kg") is not None and p.get("P_kPa")]
    d = Drawing(520, 290)
    d.add(Rect(0, 0, 520, 290, rx=14, ry=14, fillColor=colors.HexColor("#f8fafc"), strokeColor=colors.HexColor("#cbd5e1")))
    if not vals:
        d.add(String(180, 145, "Data P-h belum tersedia", fontSize=12)); return d
    hmin=min(v[0] for v in vals)-25; hmax=max(v[0] for v in vals)+30
    pmin=max(50,min(v[1] for v in vals)*0.75); pmax=max(v[1] for v in vals)*1.25
    lmin=math.log10(pmin); lmax=math.log10(pmax)
    x0,y0,w,h = 55,45,420,205
    def sx(hv): return x0 + (float(hv)-hmin)/max(hmax-hmin,1e-9)*w
    def sy(pv): return y0 + (math.log10(max(float(pv),1))-lmin)/max(lmax-lmin,1e-9)*h
    d.add(Line(x0,y0,x0,y0+h,strokeColor=colors.HexColor("#64748b")))
    d.add(Line(x0,y0,x0+w,y0,strokeColor=colors.HexColor("#64748b")))
    def polyline(points, col, sw=2):
        valid=[p for p in points if p.get("h_kJ_kg") is not None and p.get("P_kPa")]
        if len(valid)<2: return
        path=Path(); path.moveTo(sx(valid[0]["h_kJ_kg"]), sy(valid[0]["P_kPa"]))
        for p in valid[1:]: path.lineTo(sx(p["h_kJ_kg"]), sy(p["P_kPa"]))
        path.strokeColor=col; path.strokeWidth=sw; path.fillColor=None; d.add(path)
    polyline(dome.get("liquid", []), colors.HexColor("#0ea5e9"), 2)
    polyline(dome.get("vapor", []), colors.HexColor("#0f766e"), 2)
    cyc=frame.get("points", [])
    if len(cyc)>1: polyline(cyc+[cyc[0]], colors.HexColor("#1d4ed8"), 3)
    for p in cyc:
        x=sx(p.get("h_kJ_kg")); y=sy(p.get("P_kPa"))
        d.add(Circle(x,y,4.5,fillColor=colors.HexColor("#1d4ed8"),strokeColor=colors.white,strokeWidth=1))
        d.add(String(x+5,y+6,str(p.get("label","")),fontSize=8,fillColor=colors.HexColor("#0f172a")))
    d.add(String(205, 20, "Entalpi, h (kJ/kg)", fontSize=9, fillColor=colors.HexColor("#334155")))
    d.add(String(8, 150, "P log", fontSize=9, fillColor=colors.HexColor("#334155")))
    d.add(String(58, 265, f"Diagram P-h ringkas | t = {_f(frame.get('time_min'),2)} menit", fontSize=10, fillColor=colors.HexColor("#0f172a")))
    return d



def _nice_ticks(min_v: float, max_v: float, count: int = 5) -> list[float]:
    if count <= 1 or not math.isfinite(min_v) or not math.isfinite(max_v):
        return [min_v]
    if abs(max_v - min_v) < 1e-9:
        return [min_v]
    return [min_v + (max_v - min_v) * i / (count - 1) for i in range(count)]


def _line_plot_drawing(
    results: dict[str, Any],
    title: str,
    keys: list[str],
    labels: list[str],
    units: str = "",
    x_axis_label: str = "Waktu simulasi (menit)",
    y_axis_label: str = "Nilai",
) -> Drawing:
    ts = results.get("time_series", []) or results.get("visualization_series", []) or []
    d = Drawing(520, 268)
    d.add(String(18, 248, title, fontSize=12, fillColor=colors.HexColor("#0f172a")))
    d.add(String(18, 233, f"Sumbu X: {x_axis_label} | Sumbu Y: {y_axis_label}", fontSize=7.3, fillColor=colors.HexColor("#64748b")))
    if not ts:
        d.add(String(200, 130, "Data waktu belum tersedia", fontSize=10, fillColor=colors.HexColor("#64748b")))
        return d
    x0, y0, w, h = 64, 52, 408, 150
    vals: list[float] = []
    for row in ts:
        for key in keys:
            v = row.get(key)
            if isinstance(v, (int, float)) and math.isfinite(float(v)):
                vals.append(float(v))
    if not vals:
        return d
    ymin, ymax = min(vals), max(vals)
    if abs(ymax - ymin) < 1e-9:
        ymin -= 1; ymax += 1
    # Untuk grafik COP, jangan terlalu zoom agar perbedaan kecil tidak terlihat berlebihan.
    if "COP Pendinginan" in title and (ymax - ymin) < 0.50:
        mid = 0.5 * (ymax + ymin)
        ymin, ymax = mid - 0.25, mid + 0.25
    # Beri sedikit ruang agar kurva tidak menempel ke batas grafik.
    pad_y = max((ymax - ymin) * 0.08, 0.05)
    ymin -= pad_y; ymax += pad_y
    times = [float(row.get("time_min", 0) or 0) for row in ts]
    tmin = min(times); tmax = max(times) if max(times) > min(times) else min(times) + 1
    def sx(t): return x0 + (float(t) - tmin) / max(tmax - tmin, 1e-9) * w
    def sy(v): return y0 + (float(v) - ymin) / max(ymax - ymin, 1e-9) * h

    # Grid, tick, dan label sumbu.
    y_ticks = _nice_ticks(ymin, ymax, 5)
    x_ticks = _nice_ticks(tmin, tmax, 5)
    for tick in y_ticks:
        y = sy(tick)
        d.add(Line(x0, y, x0+w, y, strokeColor=colors.HexColor("#e2e8f0"), strokeWidth=0.6))
        d.add(String(x0-32, y-3, f"{tick:.1f}", fontSize=6.8, fillColor=colors.HexColor("#475569")))
    for tick in x_ticks:
        x = sx(tick)
        d.add(Line(x, y0, x, y0+h, strokeColor=colors.HexColor("#eef2f7"), strokeWidth=0.5))
        d.add(String(x-8, y0-15, f"{tick:.0f}", fontSize=6.8, fillColor=colors.HexColor("#475569")))
    d.add(Line(x0, y0, x0, y0+h, strokeColor=colors.HexColor("#94a3b8"), strokeWidth=1))
    d.add(Line(x0, y0, x0+w, y0, strokeColor=colors.HexColor("#94a3b8"), strokeWidth=1))
    d.add(String(x0 + w/2 - 42, 17, x_axis_label, fontSize=8, fillColor=colors.HexColor("#334155")))

    palette = ["#2563eb", "#dc2626", "#f59e0b", "#22c55e", "#06b6d4", "#9333ea"]
    for ki, key in enumerate(keys):
        pts = []
        for row in ts:
            v = row.get(key)
            t = row.get("time_min", 0) or 0
            if isinstance(v, (int, float)) and math.isfinite(float(v)):
                pts.append((sx(t), sy(v)))
        if len(pts) < 2:
            continue
        path = Path(); path.moveTo(pts[0][0], pts[0][1])
        for x, y in pts[1:]: path.lineTo(x, y)
        path.strokeColor = colors.HexColor(palette[ki % len(palette)]); path.strokeWidth = 2; path.fillColor = None
        d.add(path)
        lx = 304 + (ki % 2)*100; ly = 244 - (ki//2)*13
        d.add(Line(lx, ly, lx+14, ly, strokeColor=colors.HexColor(palette[ki % len(palette)]), strokeWidth=2))
        d.add(String(lx+18, ly-4, labels[ki] if ki < len(labels) else key, fontSize=7, fillColor=colors.HexColor("#334155")))
    return d


def _recommended_hwst_range_drawing(results: dict[str, Any]) -> Drawing:
    """Grafik range rekomendasi suhu HWST berbasis indeks relatif heat recovery.

    Layout dibuat lebih lega agar judul, legenda, dan keterangan tidak saling menumpuk di PDF.
    """
    ts = results.get("time_series", []) or []
    sv = results.get("summary_values", {}) or {}
    cfg = results.get("config", {}) or {}
    d = Drawing(520, 305)
    d.add(String(18, 286, "Range Suhu COP Optimum Mode DSH", fontSize=12, fillColor=colors.HexColor("#0f172a")))
    d.add(String(18, 270, "Sumbu X: Suhu air HWST (°C) | Sumbu Y: Indeks efektivitas heat recovery (%)", fontSize=7.2, fillColor=colors.HexColor("#64748b")))

    # Range COP optimum dari ringkasan model. Jika belum tersedia, pakai default konseptual.
    try:
        rec_min = float(sv.get("COPMaxRangeMin_C"))
        rec_max = float(sv.get("COPMaxRangeMax_C"))
        if not (math.isfinite(rec_min) and math.isfinite(rec_max)) or rec_max <= rec_min:
            raise ValueError
    except Exception:
        rec_min, rec_max = 40.0, float(cfg.get("T_setpoint_C") or 50.0)
        if rec_max <= rec_min:
            rec_max = 50.0
    upper_max = min(rec_max + 5.0, 60.0)
    d.add(String(18, 255, f"Range optimum model: {_range_text(rec_min, rec_max, '°C')}", fontSize=7.6, fillColor=colors.HexColor("#1d4ed8")))

    x_min, x_max = 30.0, max(60.0, rec_max + 10.0)
    y_min, y_max = 0.0, 105.0
    x0, y0, w, h = 66, 78, 404, 145
    def sx(x): return x0 + (float(x) - x_min) / max(x_max - x_min, 1e-9) * w
    def sy(y): return y0 + (float(y) - y_min) / max(y_max - y_min, 1e-9) * h
    def shade(xa, xb, color_hex):
        xa, xb = max(x_min, xa), min(x_max, xb)
        if xb > xa:
            d.add(Rect(sx(xa), y0, sx(xb)-sx(xa), h, fillColor=colors.HexColor(color_hex), strokeColor=None, fillOpacity=0.22))
    shade(rec_min, rec_max, "#bbf7d0")
    shade(rec_max, upper_max, "#fde68a")
    shade(upper_max, x_max, "#fecaca")

    # Data indeks: pakai Q_HWST aktual bila tersedia, lalu dilengkapi tren konseptual agar grafik tetap terbaca sampai 60 °C.
    raw = []
    for row in ts:
        t = row.get("T_tank_mean_C")
        q = row.get("Q_HWST_kW")
        if isinstance(t, (int, float)) and isinstance(q, (int, float)) and math.isfinite(float(t)) and math.isfinite(float(q)) and float(q) > 0:
            raw.append((float(t), float(q)))
    qmax = max([q for _, q in raw], default=1.0)
    raw = sorted(raw)
    def actual_rel(temp: float) -> float | None:
        if not raw:
            return None
        nearest = min(raw, key=lambda p: abs(p[0] - temp))
        if abs(nearest[0] - temp) <= 2.75:
            return max(0.0, min(100.0, nearest[1] / max(qmax, 1e-9) * 100.0))
        return None
    points = []
    for x in [x_min + 5*i for i in range(int((x_max-x_min)/5)+1)]:
        y = actual_rel(x)
        if y is None:
            y = max(20.0, 100.0 - 1.55*(x - 30.0) - 0.024*(x - 30.0)**2)
        points.append((x, y))

    for tick in [0, 25, 50, 75, 100]:
        d.add(Line(x0, sy(tick), x0+w, sy(tick), strokeColor=colors.HexColor("#e2e8f0"), strokeWidth=0.6))
        d.add(String(x0-30, sy(tick)-3, str(tick), fontSize=7, fillColor=colors.HexColor("#475569")))
    for tick in [x_min + 5*i for i in range(int((x_max-x_min)/5)+1)]:
        d.add(Line(sx(tick), y0, sx(tick), y0+h, strokeColor=colors.HexColor("#eef2f7"), strokeWidth=0.5))
        d.add(String(sx(tick)-6, y0-16, f"{tick:.0f}", fontSize=7, fillColor=colors.HexColor("#475569")))
    d.add(Line(x0, y0, x0, y0+h, strokeColor=colors.HexColor("#94a3b8"), strokeWidth=1))
    d.add(Line(x0, y0, x0+w, y0, strokeColor=colors.HexColor("#94a3b8"), strokeWidth=1))
    path = Path(); path.moveTo(sx(points[0][0]), sy(points[0][1]))
    for x, y in points[1:]: path.lineTo(sx(x), sy(y))
    path.strokeColor = colors.HexColor("#0f2a4d"); path.strokeWidth = 2.3; path.fillColor = None
    d.add(path)
    for x, y in points:
        d.add(Circle(sx(x), sy(y), 3.1, fillColor=colors.HexColor("#0f2a4d"), strokeColor=colors.white, strokeWidth=0.8))
    d.add(String(x0 + w/2 - 46, 34, "Suhu air HWST (°C)", fontSize=8, fillColor=colors.HexColor("#334155")))

    legend = [("#bbf7d0", "Range COP optimum"), ("#fde68a", "Batas atas operasional"), ("#fecaca", "Kurang direkomendasikan")]
    legend_x = [66, 224, 382]
    for i, (col, txt) in enumerate(legend):
        x = legend_x[i]
        y = 12
        d.add(Rect(x, y, 12, 8, fillColor=colors.HexColor(col), strokeColor=None, fillOpacity=0.58))
        d.add(String(x+18, y-1, txt, fontSize=7.1, fillColor=colors.HexColor("#334155")))
    return d

def _build_coil_donut_drawing(results: dict[str, Any], value_key: str = "persentase_akhir", title: str = "Distribusi Zona Coil Akhir Simulasi", source_key: str = "coil_usage") -> Drawing:
    rows = results.get(source_key, []) or []
    d = Drawing(520, 240)
    d.add(String(18, 218, title, fontSize=11, fillColor=colors.HexColor("#0f172a")))
    # Warna zona: DSH/SH = merah, two-phase = kuning, subcool = biru.
    color_map = {"DSH": "#dc2626", "SH": "#dc2626", "TP": "#f59e0b", "SC": "#2563eb"}
    groups = ["HWST", "Kondensor", "Evaporator"] if any(row.get("komponen") == "HWST" for row in rows) else ["Kondensor", "Evaporator"]
    def wedge(cx, cy, r, a0, a1, col):
        pts = [(cx, cy)]
        steps = max(4, int(abs(a1-a0)/10))
        for i in range(steps+1):
            a = math.radians(a0 + (a1-a0)*i/steps)
            pts.append((cx + r*math.cos(a), cy + r*math.sin(a)))
        path = Path(); path.moveTo(pts[0][0], pts[0][1])
        for x,y in pts[1:]: path.lineTo(x,y)
        path.closePath()
        path.fillColor = colors.HexColor(col); path.strokeColor = colors.white; path.strokeWidth = 0.8
        d.add(path)
    for gi, g in enumerate(groups):
        cx, cy = 92 + gi*170, 128
        r, rin = 52, 26
        gr = [row for row in rows if row.get("komponen") == g]
        angle = 90
        total = sum(float(row.get(value_key) or 0) for row in gr) or 100.0
        for row in gr:
            val = float(row.get(value_key) or 0)
            if val <= 0:
                continue
            a1 = angle - 360*val/total
            wedge(cx, cy, r, angle, a1, color_map.get(str(row.get("zona")), "#64748b"))
            angle = a1
        d.add(Circle(cx, cy, rin, fillColor=colors.white, strokeColor=colors.HexColor("#e2e8f0"), strokeWidth=1))
        d.add(String(cx-22, cy-4, g, fontSize=9, fillColor=colors.HexColor("#0f172a")))
        desc = ", ".join([f"{row.get('zona')} {float(row.get(value_key) or 0):.1f}%" for row in gr])
        d.add(String(cx-45, 50, desc[:42], fontSize=7, fillColor=colors.HexColor("#334155")))
    d.add(String(18, 18, "Keterangan: merah = DSH/SH, kuning = two-phase, biru = subcool", fontSize=7, fillColor=colors.HexColor("#64748b")))
    return d


def _cfg_num(cfg: dict[str, Any], key: str, default: float = 0.0) -> float:
    try:
        v = cfg.get(key, default)
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def _geom_value(results: dict[str, Any], component: str, parameter: str) -> Any:
    for row in results.get("geometry_rows", []) or []:
        if str(row.get("komponen")) == component and str(row.get("parameter")) == parameter:
            return row.get("value")
    return None


def _range_text(min_v: Any, max_v: Any, unit: str = "°C") -> str:
    try:
        a = float(min_v)
        b = float(max_v)
        if abs(a - b) < 1e-6:
            return f"{a:.1f} {unit}"
        return f"{a:.1f} - {b:.1f} {unit}"
    except Exception:
        return "-"


def _table(data: list[list[Any]], col_widths: list[float], font_size: int = 8) -> Table:
    tab = Table([[str(c) for c in row] for row in data], colWidths=col_widths, repeatRows=1)
    tab.setStyle(_basic_table_style(font_size=font_size))
    return tab


def _build_cop_classification_drawing(results: dict[str, Any]) -> Drawing:
    sv = results.get("summary_values", {}) or {}
    d = Drawing(520, 118)
    d.add(String(18, 96, "Klasifikasi Efisiensi Berdasarkan COP Pendinginan", fontSize=11, fillColor=colors.HexColor("#0f172a")))
    grades = ["A", "B", "C", "D", "E", "F", "G"]
    grade_colors = ["#16a34a", "#65a30d", "#a3e635", "#facc15", "#fb923c", "#f97316", "#dc2626"]
    x0, y0, w, h = 18, 58, 480, 23
    seg = w / len(grades)
    for i, (g, col) in enumerate(zip(grades, grade_colors)):
        x = x0 + i * seg
        d.add(Rect(x, y0, seg, h, fillColor=colors.HexColor(col), strokeColor=colors.white, strokeWidth=0.6))
        d.add(String(x + seg/2 - 4, y0 + 7, g, fontSize=9, fillColor=colors.white))
    d.add(String(20, 43, "A: COP > 3.60    B: >3.40    C: >3.20    D: >2.80    E: >2.60    F: >2.40    G: <=2.40", fontSize=7, fillColor=colors.HexColor("#334155")))
    markers = [
        ("AC + HWST", sv.get("COP_class_integrated"), "#1d4ed8"),
        ("Konvensional", sv.get("COP_class_conventional"), "#64748b"),
        ("Nameplate", sv.get("COP_class_nameplate"), "#0f766e"),
    ]
    for j, (label, grade, col) in enumerate(markers):
        try:
            idx = grades.index(str(grade))
            cx = x0 + idx * seg + seg / 2
        except Exception:
            cx = x0 + w + 10
        cy = 24 - j * 8
        d.add(Circle(cx, cy, 4, fillColor=colors.HexColor(col), strokeColor=colors.white, strokeWidth=0.6))
        d.add(String(cx + 8, cy - 3, f"{label}: {grade or '-'}", fontSize=7, fillColor=colors.HexColor("#334155")))
    return d


def _results_with_conventional_cop_series(results: dict[str, Any]) -> dict[str, Any]:
    sv = results.get("summary_values", {}) or {}
    baseline_cop = sv.get("COP_AC_conventional")
    cloned = dict(results)
    ts = []
    for row in results.get("time_series", []) or results.get("visualization_series", []) or []:
        nr = dict(row)
        nr["COP_AC_conventional_line"] = baseline_cop
        ts.append(nr)
    cloned["time_series"] = ts
    return cloned


def _build_conclusion(results: dict[str, Any]) -> str:
    sv = results.get("summary_values", {}) or {}
    cfg = results.get("config", {}) or {}
    return (
        f"Berdasarkan hasil simulasi, penambahan HWST pada sistem AC split mampu memanfaatkan panas buang refrigeran "
        f"untuk memanaskan air hingga mencapai set point {_f(cfg.get('T_setpoint_C'), 1)} °C dalam waktu {_f(sv.get('reachTime_min'), 2)} menit. "
        f"Nilai COP AC + HWST sebesar {_f(sv.get('COP_AC_integrated'), 3)}, COP AC konvensional sebesar {_f(sv.get('COP_AC_conventional'), 3)}, "
        f"dan COP useful sebesar {_f(sv.get('COP_useful_integrated'), 3)} menunjukkan bahwa sistem tidak hanya menghasilkan efek pendinginan, "
        f"tetapi juga memberikan manfaat tambahan berupa pemanasan air dari energi panas yang sebelumnya terbuang ke lingkungan."
    )


def build_pdf(results: dict[str, Any]) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=34, leftMargin=34, topMargin=34, bottomMargin=34)
    styles = getSampleStyleSheet()
    story: list[Any] = []

    cfg = results.get("config", {}) or {}
    sv = results.get("summary_values", {}) or {}
    engine = results.get("engine_info", {}) or {}
    capacity_kW = _cfg_num(cfg, "ac_capacity_pk") * _cfg_num(cfg, "cooling_capacity_per_PK_kW")
    capacity_BTUh = capacity_kW * 3412.142

    story.append(Paragraph("Report Hasil Simulasi AC + HWST", styles["Title"]))
    story.append(Paragraph(f"Ringkasan performa sistem berdasarkan hasil simulasi web. Versi model: {engine.get('engine_version', 'V17.4')}.", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("1. Data Input Utama", styles["Heading2"]))
    main_data = [
        ["Parameter", "Nilai", "Unit"],
        ["Refrigerant", cfg.get("refrigerant", "-"), "-"],
        ["Kapasitas pendinginan", f"{capacity_BTUh:.0f} / {capacity_kW:.3f}", "Btu/h / kW"],
        ["Daya kompresor", _f(cfg.get("compressor_power_kW"), 3), "kW"],
        ["Volume tangki", _f(cfg.get("tank_volume_L"), 1), "L"],
        ["Suhu awal air", _f(cfg.get("T_tank_initial_C"), 1), "°C"],
        ["Set point air", _f(cfg.get("T_setpoint_C"), 1), "°C"],
        ["Metode coil", engine.get("coil_method", "NTU-effectiveness berbasis zona"), "-"],
    ]
    story.append(_table(main_data, [190, 170, 90], 8))
    story.append(Spacer(1, 12))

    story.append(Paragraph("2. Luas Efektif dan Nilai K Coil", styles["Heading2"]))
    coil_data = [
        ["Komponen", "Luas efektif coil", "Nilai K tube", "Unit K"],
        ["HWST", _f(_geom_value(results, "HWST", "Area tube"), 4), _f(cfg.get("k_tube_hwst_W_mK"), 2), "W/m.K"],
        ["Kondensor", _f(_geom_value(results, "Kondensor", "Area total efektif"), 4), _f(cfg.get("k_tube_cond_W_mK"), 2), "W/m.K"],
        ["Evaporator", _f(_geom_value(results, "Evaporator", "Area total efektif"), 4), _f(cfg.get("k_tube_evap_W_mK"), 2), "W/m.K"],
    ]
    story.append(_table(coil_data, [110, 140, 120, 80], 8))
    story.append(Spacer(1, 12))

    story.append(Paragraph("3. Hasil Utama Simulasi", styles["Heading2"]))
    results_data = [
        ["Output", "Nilai", "Unit"],
        ["Waktu mencapai set point", _f(sv.get("reachTime_min"), 2), "menit"],
        ["Suhu akhir tangki", _f(sv.get("finalTankMean_C"), 2), "°C"],
        ["Energi HWST", _f(sv.get("energy_HWST_kWh"), 4), "kWh"],
        ["COP AC + HWST", _f(sv.get("COP_AC_integrated"), 3), "-"],
        ["COP AC konvensional", _f(sv.get("COP_AC_conventional"), 3), "-"],
        ["Peningkatan COP vs konvensional", f"{_f(sv.get('delta_COP_integrated'), 3)} / {_f(sv.get('delta_COP_integrated_pct'), 2)}%", "-"],
        ["COP useful", _f(sv.get("COP_useful_integrated"), 3), "-"],
        ["Range suhu COP optimum mode DSH", _range_text(sv.get("COPMaxRangeMin_C"), sv.get("COPMaxRangeMax_C")), "-"],
    ]
    story.append(_table(results_data, [210, 130, 80], 8))
    story.append(Spacer(1, 8))
    condition_rows = results.get("condition_rows", []) or []
    if condition_rows:
        cond_table = [["Sistem", "Parameter", "Nilai", "Unit", "Catatan"]]
        for r in condition_rows[:4]:
            cond_table.append([r.get("sistem"), r.get("parameter"), r.get("nilai"), r.get("unit"), r.get("catatan")])
        story.append(Paragraph("Kondisi Refrigeran Akhir", styles["Heading3"]))
        story.append(_table(cond_table, [82, 150, 60, 38, 120], 7))
        story.append(Spacer(1, 8))
    pd_rows = results.get("pressure_drop_rows", []) or []
    if pd_rows:
        pd_table = [["Sistem", "Komponen", "Nilai", "Unit"]]
        for r in pd_rows:
            if r.get("parameter") in {"Pressure drop", "DP total"}:
                pd_table.append([r.get("sistem"), r.get("komponen"), r.get("nilai"), r.get("unit")])
        story.append(Paragraph("Ringkasan Pressure Drop", styles["Heading3"]))
        story.append(_table(pd_table[:12], [115, 100, 75, 45], 7))
        story.append(Spacer(1, 12))

    story.append(Paragraph("4. Klasifikasi COP", styles["Heading2"]))
    class_data = [
        ["Jenis COP", "Nilai", "Kelas", "Rentang"],
        ["COP AC + HWST", _f(sv.get("COP_AC_integrated"), 3), sv.get("COP_class_integrated", "-"), sv.get("COP_class_integrated_range", "-")],
        ["COP AC konvensional", _f(sv.get("COP_AC_conventional"), 3), sv.get("COP_class_conventional", "-"), sv.get("COP_class_conventional_range", "-")],
        ["COP nameplate", _f(sv.get("COP_nameplate"), 3), sv.get("COP_class_nameplate", "-"), sv.get("COP_class_nameplate_range", "-")],
    ]
    story.append(_table(class_data, [145, 75, 55, 160], 8))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Catatan: COP useful tidak diklasifikasikan karena memasukkan manfaat pemanasan air, sedangkan klasifikasi efisiensi AC mengacu pada COP pendinginan.", styles["Normal"]))
    story.append(Spacer(1, 8))
    story.append(_build_cop_classification_drawing(results))

    story.append(PageBreak())
    story.append(Paragraph("5. Persentase Zona Coil", styles["Heading2"]))
    story.append(_build_coil_donut_drawing(results, value_key="persentase_akhir", title="AC + HWST — Persentase Zona Coil Akhir Simulasi", source_key="coil_usage"))
    if results.get("coil_usage_conventional"):
        story.append(Spacer(1, 8))
        story.append(_build_coil_donut_drawing(results, value_key="persentase_steady", title="AC Konvensional — Persentase Zona Coil Steady", source_key="coil_usage_conventional"))
    story.append(PageBreak())
    story.append(KeepTogether([
        Paragraph("6. Grafik Hasil Simulasi", styles["Heading2"]),
        Spacer(1, 6),
        _recommended_hwst_range_drawing(results),
    ]))
    story.append(Spacer(1, 10))
    story.append(_line_plot_drawing(results, "Temperatur Air Tangki terhadap Waktu", ["T_tank_mean_C"], ["T air tangki"], "°C", "Waktu simulasi (menit)", "Temperatur air tangki (°C)"))
    story.append(Spacer(1, 10))
    cop_results = _results_with_conventional_cop_series(results)
    story.append(_line_plot_drawing(cop_results, "COP Pendinginan terhadap Waktu", ["COP_AC_conventional_line", "COP_AC"], ["COP konvensional", "COP AC + HWST"], "", "Waktu simulasi (menit)", "COP pendinginan"))
    story.append(Spacer(1, 10))
    story.append(_line_plot_drawing(results, "COP Useful terhadap Waktu", ["COP_useful"], ["COP useful"], "", "Waktu simulasi (menit)", "COP useful"))

    story.append(Spacer(1, 14))
    story.append(Paragraph("7. Kesimpulan", styles["Heading2"]))
    story.append(Paragraph(_build_conclusion(results), styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer

